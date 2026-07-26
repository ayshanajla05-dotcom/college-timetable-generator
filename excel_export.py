from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font

from models import Timetable


def generate_excel():

    wb = Workbook()
    ws = wb.active
    ws.title = "College Timetable"

    headers = [
        "Section",
        "Day Order",
        "Period",
        "Subject",
        "Faculty",
        "Room"
    ]

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    timetable = Timetable.query.order_by(
        Timetable.section_id,
        Timetable.day_order,
        Timetable.period
    ).all()

    row_no = 2

    for row in timetable:

        ws.cell(row=row_no, column=1).value = row.section.section_name
        ws.cell(row=row_no, column=2).value = row.day_order
        ws.cell(row=row_no, column=3).value = row.period
        ws.cell(row=row_no, column=4).value = row.subject.subject_name
        ws.cell(row=row_no, column=5).value = row.faculty.faculty_name
        ws.cell(row=row_no, column=6).value = row.room.room_number

        row_no += 1

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="College_Timetable.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )